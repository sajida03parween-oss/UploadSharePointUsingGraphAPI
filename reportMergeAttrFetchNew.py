import csv
import os
import logging
from logging.handlers import RotatingFileHandler
from datetime import datetime

import pyodbc
from openpyxl import load_workbook   # NEW: use openpyxl

# ======================================================
# TEST CONFIG
# ======================================================

TEST_MODE = False
TEST_LIMIT = 10000

# =====================================================
# DEFINE LOOKUP TABLE
# =====================================================

def load_lookup_table(conn, table_name, key_col, value_col):
    query = f"""
    SELECT {key_col}, {value_col}
    FROM dbo.{table_name}
    """
    cursor = conn.cursor()
    cursor.execute(query)
    lookup = {row[0]: row[1] for row in cursor.fetchall()}
    logger.info(f"Loaded lookup table: {table_name} | Rows: {len(lookup)}")
    return lookup

# ======================================================
# LOOKUP MAPPING
# ======================================================
LOOKUP_CONFIG = {
    "CN_DOCUMENT_APPLICABILITY": {
        "table": "TN_DOCUMENT_APPLICABILITY",
        "key": "OBJECT_ID",
        "value": "TDM_NAME"
    },
    "CN_DOC_LIST_1": {
        "table": "TN_DOC_LIST_1",
        "key": "OBJECT_ID",
        "value": "TDM_NAME"
    }
}

# =========================================================
# CONFIGURATION
# =========================================================

SERVER = r"YS00583Q\SQLEXPRESS"
DATABASE = "Copy_RECT"

CONNECTION_STRING = (
     "DRIVER={ODBC Driver 18 for SQL Server};"
     f"SERVER={SERVER};"
     f"DATABASE={DATABASE};"
      "UID=sa;"
     "PWD=R7!vQ9#Zx@2L$A8K;"
     "TrustServerCertificate=yes;"
 )

EXPORT_FOLDER = r"D:\Utility_Scripts\Pavan"
LOG_FOLDER = r"D:\Utility_Scripts\Pavan\logs"

TABLE_CONFIG = {
    "TN_DOCUMENTS": [
        "OBJECT_ID",
        "CLASS_ID",
        "TDMX_ID",
        "CN_DOCUMENT_APPLICABILITY",
        "CN_DOC_LIST_1",
        "TDMX_DETAILED_DESCRIPTION",
        "TDMX_COMMENTS",
        "TDMX_SW_CONFIGURATION"
    ]
}

# ======================================================
# CREATE FOLDERS
# ======================================================

os.makedirs(EXPORT_FOLDER, exist_ok=True)
os.makedirs(LOG_FOLDER, exist_ok=True)

# ======================================================
# LOGGING
# ======================================================

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
log_file = os.path.join(LOG_FOLDER, f"export_log_{timestamp}.txt")

logger = logging.getLogger("ExportLogger")
logger.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

file_handler = RotatingFileHandler(log_file, maxBytes=10 * 1024 * 1024, backupCount=5)
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

# ======================================================
# DATABASE CONNECTION
# ======================================================

def get_connection():
    logger.info("Connecting to database")
    conn = pyodbc.connect(CONNECTION_STRING)
    logger.info("Database connected")
    return conn

# ======================================================
# READ EXCEL INPUT
# ======================================================

def read_input_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    input_data = []
    # Assuming first row is header: TDMX_ID | Revision
    for row in ws.iter_rows(min_row=2, values_only=True):
        tdmx_id, revision = row
        if tdmx_id is not None and revision is not None:
            input_data.append((tdmx_id, revision))
    logger.info(f"Loaded {len(input_data)} rows from Excel input")
    return input_data

# ======================================================
# EXPORT FUNCTION
# ======================================================

def export_table(conn, table_name, columns, lookup_cache, input_data):
    logger.info(f"Starting export for {table_name}")
    try:
        csv_file = os.path.join(EXPORT_FOLDER, f"{table_name}_{timestamp}.csv")
        cursor = conn.cursor()
        rows_exported = 0

        with open(csv_file, mode="w", newline="", encoding="utf-8-sig") as file:
            writer = csv.writer(file)
            writer.writerow(columns)  # CSV header

            for tdmx_id, revision in input_data:
                query = f"""
                SELECT {", ".join(columns)}
                FROM dbo.{table_name}
                WHERE TDMX_ID = ? AND Revision = ?
                """
                cursor.execute(query, (tdmx_id, revision))
                rows = cursor.fetchall()

                processed_rows = []
                for row in rows:
                    row = list(row)
                    for idx, column_name in enumerate(columns):
                        if column_name in LOOKUP_CONFIG:
                            lookup_value = lookup_cache[column_name].get(row[idx])
                            row[idx] = lookup_value if lookup_value else ""

                    try:
                        applicability_idx = columns.index("CN_DOCUMENT_APPLICABILITY")
                        detailed_desc_idx = columns.index("TDMX_DETAILED_DESCRIPTION")
                        comments_idx = columns.index("TDMX_COMMENTS")
                        design_module_idx = columns.index("CN_DOC_LIST_1")
                        sw_config_idx = columns.index("TDMX_SW_CONFIGURATION")

                        merged_text = (
                            f"A/C Applicability: {row[applicability_idx] or ''}\n"
                            f"Design Module: {row[design_module_idx] or ''}\n"
                            f"Details: {row[detailed_desc_idx] or ''}\n"
                            f"Comments: {row[comments_idx] or ''}"
                        )
                        row[sw_config_idx] = merged_text
                    except Exception as merge_error:
                        logger.warning(f"Merge failed for row: {str(merge_error)}")

                    processed_rows.append(row)

                writer.writerows(processed_rows)
                rows_exported += len(processed_rows)
                logger.info(f"{table_name} | Exported rows: {rows_exported}")

        logger.info(f"SUCCESS | {table_name} | Total Rows: {rows_exported}")
        print(f"CSV CREATED: {csv_file}")

    except Exception as e:
        logger.exception(f"FAILED | {table_name} | {str(e)}")

# ======================================================
# MAIN
# ======================================================

def main():
    logger.info("===================================")
    logger.info("EXPORT PROCESS STARTED")
    logger.info("===================================")

    try:
        conn = get_connection()

        # Load lookup tables
        lookup_cache = {
            col: load_lookup_table(conn, cfg["table"], cfg["key"], cfg["value"])
            for col, cfg in LOOKUP_CONFIG.items()
        }

        # Read Excel input
        input_file = r"D:\Utility_Scripts\Pavan\inputAMFromReports.xlsx"
        input_data = read_input_excel(input_file)

        # Export tables
        for table_name, columns in TABLE_CONFIG.items():
            export_table(conn, table_name, columns, lookup_cache, input_data)

        conn.close()
        logger.info("Database connection closed")

    except Exception as e:
        logger.exception(f"FATAL ERROR | {str(e)}")

    logger.info("===================================")
    logger.info("EXPORT PROCESS COMPLETED")
    logger.info("===================================")

# ======================================================
# ENTRY
# ======================================================

if __name__ == "__main__":
    main()
